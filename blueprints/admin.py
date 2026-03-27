from functools import wraps
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify
from supabase_client import get_supabase
import config
from openpyxl import load_workbook
from datetime import date
import json

admin_bp = Blueprint("admin", __name__, url_prefix="/admin",
                     template_folder="../templates/admin")


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_stock_file(file_storage):
    workbook = load_workbook(file_storage, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["The Excel file is empty."]

    header_row = [_normalize_header(cell) for cell in rows[0]]
    data_rows = rows[1:]

    name_headers = {"item_name", "name", "item", "itemcode", "item_code", "code"}
    qty_headers = {"stock", "qty", "quantity", "stock_qty", "stock_quantity"}

    name_idx = next((idx for idx, value in enumerate(header_row) if value in name_headers), None)
    qty_idx = next((idx for idx, value in enumerate(header_row) if value in qty_headers), None)

    if name_idx is None or qty_idx is None:
        if len(rows[0]) >= 2:
            name_idx = 0
            qty_idx = 1
            data_rows = rows
        else:
            return [], ["Could not detect columns. Use headers like item_name and stock."]

    parsed_rows = []
    issues = []
    for row_number, row in enumerate(data_rows, start=2 if data_rows is rows[1:] else 1):
        if row is None:
            continue

        raw_name = row[name_idx] if name_idx < len(row) else None
        raw_qty = row[qty_idx] if qty_idx < len(row) else None

        if raw_name in (None, "") and raw_qty in (None, ""):
            continue

        item_name = str(raw_name or "").strip()
        if not item_name:
            issues.append(f"Row {row_number}: missing item name.")
            continue

        try:
            quantity = int(float(raw_qty))
        except (TypeError, ValueError):
            issues.append(f"Row {row_number}: invalid stock value for {item_name}.")
            continue

        if quantity == 0:
            issues.append(f"Row {row_number}: stock cannot be 0 for {item_name}.")
            continue

        parsed_rows.append({"item_name": item_name, "quantity": quantity})

    if not parsed_rows and not issues:
        issues.append("No valid rows found in the Excel file.")

    return parsed_rows, issues


def _add_stock_quantity(sb, warehouse_id, item_id, quantity):
    existing = sb.table("warehouse_stock").select("id, stock").eq(
        "warehouse_id", warehouse_id
    ).eq("item_id", item_id).execute().data

    if existing:
        new_stock = int(existing[0]["stock"]) + int(quantity)
        sb.table("warehouse_stock").update({"stock": new_stock}).eq(
            "id", existing[0]["id"]
        ).execute()
    else:
        sb.table("warehouse_stock").insert({
            "warehouse_id": warehouse_id,
            "item_id": item_id,
            "stock": int(quantity),
        }).execute()


# ---------- Auth helpers ----------
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("user.login"))
        return f(*args, **kwargs)
    return decorated


# ---------- Login / Logout ----------
@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        return redirect(url_for("user.login"), code=307)
    return redirect(url_for("user.login"))


@admin_bp.route("/logout")
def logout():
    session.pop("is_admin", None)
    flash("Logged out.", "info")
    return redirect(url_for("user.login"))


def _update_pending_invoice_line_items(sb, inv_id, form_data):
    line_items = sb.table("invoice_items").select("id, item_id, quantity, discount").eq("invoice_id", inv_id).execute().data

    party_raw = form_data.get("party_id", "").strip()
    adda_raw = form_data.get("adda_id", "").strip()
    invoice_date_raw = form_data.get("invoice_date", "").strip()
    delivery_paid_raw = form_data.get("delivery_paid", "yes").strip().lower()
    delivery_amount_raw = form_data.get("delivery_amount", "0").strip() or "0"

    try:
        party_id = int(party_raw)
        adda_id = int(adda_raw)
    except ValueError:
        return False, "Please select a valid party and adda."

    if party_id <= 0 or adda_id <= 0:
        return False, "Please select a valid party and adda."

    if not invoice_date_raw:
        return False, "Invoice date is required."

    try:
        date.fromisoformat(invoice_date_raw)
    except ValueError:
        return False, "Invalid invoice date format."

    delivery_paid = delivery_paid_raw == "yes"
    try:
        delivery_amount = float(delivery_amount_raw)
    except ValueError:
        return False, "Invalid delivery amount."

    if delivery_amount < 0:
        return False, "Delivery amount cannot be negative."

    if delivery_paid:
        delivery_amount = 0.0

    delete_item_ids = set()
    for raw_id in form_data.getlist("delete_item_ids"):
        raw_id = (raw_id or "").strip()
        if not raw_id:
            continue
        try:
            delete_item_ids.add(int(raw_id))
        except ValueError:
            return False, "Invalid item delete request."

    remaining_items = [li for li in line_items if int(li["id"]) not in delete_item_ids]

    new_items_raw = form_data.get("new_items_json", "[]").strip() or "[]"
    try:
        new_items = json.loads(new_items_raw)
    except (TypeError, ValueError, json.JSONDecodeError):
        return False, "Invalid new items payload."

    if not isinstance(new_items, list):
        return False, "Invalid new items payload."

    sanitized_new_items = []
    for entry in new_items:
        if not isinstance(entry, dict):
            return False, "Invalid new item entry."

        item_raw = str(entry.get("item_id", "")).strip()
        qty_raw = str(entry.get("quantity", "")).strip()
        disc_raw = str(entry.get("discount", "0")).strip() or "0"

        try:
            item_id = int(item_raw)
            quantity = int(qty_raw)
        except ValueError:
            return False, "New item must have valid item and quantity."

        try:
            line_discount = float(disc_raw)
        except ValueError:
            line_discount = 0.0

        if item_id <= 0 or quantity <= 0:
            return False, "New item must have valid item and quantity."

        sanitized_new_items.append({
            "item_id": item_id,
            "quantity": quantity,
            "discount": max(0.0, line_discount),
        })

    if not remaining_items and not sanitized_new_items:
        return False, "Invoice must have at least one item."

    updates = []
    for li in remaining_items:
        qty_raw = form_data.get(f"item_quantity_{li['id']}", str(li.get("quantity", 0))).strip() or "0"
        disc_raw = form_data.get(f"item_discount_{li['id']}", str(li.get("discount", 0))).strip() or "0"

        try:
            quantity = int(qty_raw)
        except ValueError:
            return False, "Invalid quantity value provided."

        try:
            line_discount = float(disc_raw)
        except ValueError:
            line_discount = 0.0

        if quantity <= 0:
            return False, "Quantity must be greater than 0 for every invoice item."

        updates.append({
            "id": int(li["id"]),
            "quantity": quantity,
            "discount": max(0.0, line_discount),
        })

    sb.table("invoices").update({
        "party_id": party_id,
        "adda_id": adda_id,
        "invoice_date": invoice_date_raw,
        "delivery_paid": delivery_paid,
        "delivery_amount": delivery_amount,
    }).eq("id", inv_id).execute()

    if delete_item_ids:
        sb.table("invoice_items").delete().eq("invoice_id", inv_id).in_("id", list(delete_item_ids)).execute()

    for upd in updates:
        sb.table("invoice_items").update({
            "quantity": upd["quantity"],
            "discount": upd["discount"],
        }).eq("id", upd["id"]).execute()

    if sanitized_new_items:
        sb.table("invoice_items").insert([
            {
                "invoice_id": inv_id,
                "item_id": it["item_id"],
                "quantity": it["quantity"],
                "discount": it["discount"],
            }
            for it in sanitized_new_items
        ]).execute()

    deleted_count = len(delete_item_ids)
    added_count = len(sanitized_new_items)
    parts = ["Invoice updated."]
    if deleted_count:
        parts.append(f"{deleted_count} item(s) deleted.")
    if added_count:
        parts.append(f"{added_count} item(s) added.")
    return True, " ".join(parts)


# ---------- Dashboard ----------
@admin_bp.route("/")
@admin_required
def dashboard():
    sb = get_supabase()
    parties = sb.table("parties").select("*").order("name").execute().data
    items = sb.table("items").select("*").order("name").execute().data
    addas = sb.table("addas").select("*").order("name").execute().data
    warehouses = sb.table("warehouses").select("*").order("name").execute().data
    stock = sb.table("warehouse_stock").select(
        "*, warehouses(name), items(item_code, name, box_qty)"
    ).order("warehouse_id").execute().data

    # Stock positions matrix: include every item for every warehouse (default stock = 0)
    stock_map = {
        (int(s["warehouse_id"]), int(s["item_id"])): int(s.get("stock") or 0)
        for s in stock
    }
    stock_positions = []
    stock_totals = []
    for wh in warehouses:
        wh_id = int(wh["id"])
        wh_name = wh.get("name") or "-"
        for item in items:
            item_id = int(item["id"])
            stock_positions.append({
                "warehouse_id": wh_id,
                "warehouse_name": wh_name,
                "item_id": item_id,
                "item_code": item.get("item_code") or "?",
                "item_name": item.get("name") or "-",
                "box_qty": int(item.get("box_qty") or 1),
                "stock": stock_map.get((wh_id, item_id), 0),
            })

    for item in items:
        item_id = int(item["id"])
        total_stock = sum(
            stock_map.get((int(wh["id"]), item_id), 0)
            for wh in warehouses
        )
        stock_totals.append({
            "item_id": item_id,
            "item_code": item.get("item_code") or "?",
            "item_name": item.get("name") or "-",
            "box_qty": int(item.get("box_qty") or 1),
            "stock": total_stock,
        })
    pending = sb.table("invoices").select(
        "*, parties(name), addas(name, number), invoice_items(*, items(item_code, name, box_qty, discount))"
    ).eq("status", "pending").order("created_at", desc=True).execute().data
    approved_rows = sb.table("approved_invoices").select(
        "*, parties(name), addas(name, number), warehouses(name), invoices(status), approved_invoice_items(*, items(item_code, name, box_qty, discount))"
    ).order("approved_at", desc=True).limit(50).execute().data

    approved = []
    generated = []
    for row in approved_rows:
        invoice_status = ((row.get("invoices") or {}).get("status") or "").strip()
        if invoice_status == "invoiceGenerated":
            generated.append(row)
        elif invoice_status == "approved":
            approved.append(row)

    return render_template("dashboard.html",
                           parties=parties, items=items, addas=addas,
                           warehouses=warehouses, stock=stock,
                           stock_positions=stock_positions,
                           stock_totals=stock_totals,
                           pending=pending, approved=approved, generated=generated)


# ========== PARTIES CRUD ==========
@admin_bp.route("/parties/add", methods=["POST"])
@admin_required
def add_party():
    name = request.form.get("name", "").strip()
    if name:
        get_supabase().table("parties").insert({"name": name}).execute()
        flash("Party added.", "success")
    return redirect(url_for("admin.dashboard") + "#parties")


@admin_bp.route("/parties/delete/<int:pid>")
@admin_required
def delete_party(pid):
    get_supabase().table("parties").delete().eq("id", pid).execute()
    flash("Party deleted.", "warning")
    return redirect(url_for("admin.dashboard") + "#parties")


# ========== ITEMS CRUD ==========
@admin_bp.route("/items/add", methods=["POST"])
@admin_required
def add_item():
    item_code = request.form.get("item_code", "").strip()
    name = request.form.get("name", "").strip()
    box_qty = request.form.get("box_qty", "1").strip() or "1"
    discount_raw = request.form.get("discount", "0").strip() or "0"
    try:
        discount = float(discount_raw)
    except ValueError:
        discount = 0.0
    discount = max(0.0, discount)
    if item_code and name:
        get_supabase().table("items").insert({
            "item_code": item_code,
            "name": name,
            "box_qty": max(1, int(box_qty)),
            "discount": discount,
        }).execute()
        flash("Item added.", "success")
    return redirect(url_for("admin.dashboard") + "#items")


@admin_bp.route("/items/update/<int:iid>", methods=["POST"])
@admin_required
def update_item(iid):
    discount_raw = request.form.get("discount", "0").strip() or "0"
    try:
        discount = float(discount_raw)
    except ValueError:
        discount = 0.0
    discount = max(0.0, discount)

    get_supabase().table("items").update({"discount": discount}).eq("id", iid).execute()
    flash("Item updated.", "success")
    return redirect(url_for("admin.dashboard") + "#items")


@admin_bp.route("/items/delete/<int:iid>")
@admin_required
def delete_item(iid):
    get_supabase().table("items").delete().eq("id", iid).execute()
    flash("Item deleted.", "warning")
    return redirect(url_for("admin.dashboard") + "#items")


# ========== ADDAS CRUD ==========
@admin_bp.route("/addas/add", methods=["POST"])
@admin_required
def add_adda():
    name = request.form.get("name", "").strip()
    number = request.form.get("number", "").strip()
    if name:
        get_supabase().table("addas").insert({"name": name, "number": number}).execute()
        flash("Adda added.", "success")
    return redirect(url_for("admin.dashboard") + "#addas")


@admin_bp.route("/addas/delete/<int:aid>")
@admin_required
def delete_adda(aid):
    get_supabase().table("addas").delete().eq("id", aid).execute()
    flash("Adda deleted.", "warning")
    return redirect(url_for("admin.dashboard") + "#addas")


@admin_bp.route("/addas/rename/<int:aid>", methods=["POST"])
@admin_required
def rename_adda(aid):
    new_name = request.form.get("name", "").strip()
    new_number = request.form.get("number", "").strip()
    if new_name:
        get_supabase().table("addas").update({"name": new_name, "number": new_number}).eq("id", aid).execute()
        flash("Adda renamed.", "success")
    return redirect(url_for("admin.dashboard") + "#addas")


# ========== INVOICE REVIEW (view full invoice + pick warehouse + approve) ==========
@admin_bp.route("/invoices/review/<int:inv_id>")
@admin_required
def review_invoice(inv_id):
    sb = get_supabase()
    rows = sb.table("invoices").select(
        "*, parties(name), addas(name, number), invoice_items(*, items(item_code, name, box_qty, discount))"
    ).eq("id", inv_id).execute().data
    if not rows:
        flash("Invoice not found.", "danger")
        return redirect(url_for("admin.dashboard") + "#pending")
    inv = rows[0]
    parties = sb.table("parties").select("id, name, addr").order("name").execute().data
    addas = sb.table("addas").select("id, name, number").order("name").execute().data
    warehouses = sb.table("warehouses").select("*").order("name").execute().data

    # Fetch per-warehouse stock for each item in the invoice table view,
    # and build warehouse stock map for all items for client-side checks.
    item_ids = [li["item_id"] for li in inv.get("invoice_items", [])]
    item_stock_map = {}  # {item_id: [{warehouse_name, warehouse_id, stock}, ...]}
    # Also build {warehouse_id: {item_id: stock}} for JS validation
    wh_stock_json = {}  # {str(warehouse_id): {str(item_id): stock}}
    stock_rows = sb.table("warehouse_stock").select(
        "item_id, warehouse_id, stock, warehouses(name)"
    ).execute().data
    item_id_set = set(item_ids)
    for sr in stock_rows:
        iid = sr["item_id"]
        wid = sr["warehouse_id"]
        wh_name = sr["warehouses"]["name"] if sr.get("warehouses") else "Unknown"

        if iid in item_id_set:
            if iid not in item_stock_map:
                item_stock_map[iid] = []
            item_stock_map[iid].append({"warehouse": wh_name, "warehouse_id": wid, "stock": sr["stock"]})

        wid_s = str(wid)
        if wid_s not in wh_stock_json:
            wh_stock_json[wid_s] = {}
        wh_stock_json[wid_s][str(iid)] = sr["stock"]

    # Build invoice items list for JS: [{item_id, item_code, name, quantity}, ...]
    inv_items_json = [
        {"line_id": li["id"],
         "item_id": li["item_id"],
         "item_code": li["items"]["item_code"] if li.get("items") else "?",
         "name": li["items"]["name"] if li.get("items") else "-",
         "quantity": li["quantity"]}
        for li in inv.get("invoice_items", [])
    ]

    return render_template("review_invoice.html", inv=inv, parties=parties, addas=addas, warehouses=warehouses,
                           item_stock_map=item_stock_map,
                           wh_stock_json=wh_stock_json,
                           inv_items_json=inv_items_json)


# ========== INVOICE APPROVAL (POST with warehouse) ==========
@admin_bp.route("/invoices/approve/<int:inv_id>", methods=["POST"])
@admin_required
def approve_invoice(inv_id):
    sb = get_supabase()
    warehouse_id = request.form.get("warehouse_id")
    if not warehouse_id:
        flash("Please select a warehouse.", "danger")
        return redirect(url_for("admin.review_invoice", inv_id=inv_id))

    warehouse_id = int(warehouse_id)

    rows = sb.table("invoices").select("id").eq("id", inv_id).eq("status", "pending").execute().data
    if not rows:
        flash("Invoice not found or already processed.", "danger")
        return redirect(url_for("admin.dashboard") + "#pending")

    ok, err = _update_pending_invoice_line_items(sb, inv_id, request.form)
    if not ok:
        flash(err, "danger")
        return redirect(url_for("admin.review_invoice", inv_id=inv_id))

    # Fetch updated pending invoice with line items
    rows = sb.table("invoices").select("*").eq("id", inv_id).eq("status", "pending").execute().data
    if not rows:
        flash("Invoice not found or already processed.", "danger")
        return redirect(url_for("admin.dashboard") + "#pending")
    inv = rows[0]

    # Copy header to approved_invoices (with warehouse)
    approved_result = sb.table("approved_invoices").insert({
        "invoice_id": inv["id"],
        "invoice_number": inv["invoice_number"],
        "party_id": inv["party_id"],
        "adda_id": inv["adda_id"],
        "warehouse_id": warehouse_id,
        "delivery_paid": inv["delivery_paid"],
        "delivery_amount": inv["delivery_amount"],
        "invoice_date": inv["invoice_date"],
    }).execute()
    approved_id = approved_result.data[0]["id"]

    # Copy line items to approved_invoice_items
    line_items = sb.table("invoice_items").select("*").eq("invoice_id", inv_id).execute().data
    if line_items:
        approved_lines = [
            {
                "approved_invoice_id": approved_id,
                "item_id": li["item_id"],
                "quantity": li["quantity"],
                "discount": li.get("discount", 0),
            }
            for li in line_items
        ]
        sb.table("approved_invoice_items").insert(approved_lines).execute()

        # Deduct stock from chosen warehouse
        # If a warehouse/item stock row does not exist, create it with negative stock.
        qty_by_item = {}
        for li in line_items:
            item_id = int(li["item_id"])
            qty_by_item[item_id] = qty_by_item.get(item_id, 0) + int(li["quantity"])

        if qty_by_item:
            stock_rows = sb.table("warehouse_stock").select("id, item_id, stock").eq(
                "warehouse_id", warehouse_id
            ).in_("item_id", list(qty_by_item.keys())).execute().data

            stock_map = {int(r["item_id"]): r for r in stock_rows}
            for item_id, qty in qty_by_item.items():
                existing = stock_map.get(item_id)
                if existing:
                    new_stock = int(existing["stock"]) - qty
                    sb.table("warehouse_stock").update({"stock": new_stock}).eq(
                        "id", existing["id"]
                    ).execute()
                else:
                    sb.table("warehouse_stock").insert({
                        "warehouse_id": warehouse_id,
                        "item_id": item_id,
                        "stock": -qty,
                    }).execute()

    # Update status
    sb.table("invoices").update({"status": "approved"}).eq("id", inv_id).execute()
    flash("Invoice approved! Stock deducted from warehouse.", "success")
    return redirect(url_for("admin.dashboard") + "#pending")


@admin_bp.route("/invoices/update/<int:inv_id>", methods=["POST"])
@admin_required
def update_invoice(inv_id):
    sb = get_supabase()
    rows = sb.table("invoices").select("id, status").eq("id", inv_id).execute().data
    if not rows:
        flash("Invoice not found.", "danger")
        return redirect(url_for("admin.dashboard") + "#pending")

    if rows[0].get("status") != "pending":
        flash("Only pending invoices can be edited.", "warning")
        return redirect(url_for("admin.review_invoice", inv_id=inv_id))

    ok, msg = _update_pending_invoice_line_items(sb, inv_id, request.form)
    if not ok:
        flash(msg, "danger")
    else:
        flash("Invoice changes saved.", "success")
    return redirect(url_for("admin.review_invoice", inv_id=inv_id))


@admin_bp.route("/invoices/reject/<int:inv_id>")
@admin_required
def reject_invoice(inv_id):
    get_supabase().table("invoices").update({"status": "rejected"}).eq("id", inv_id).execute()
    flash("Invoice rejected.", "warning")
    return redirect(url_for("admin.dashboard") + "#pending")


@admin_bp.route("/invoices/delete/<int:inv_id>")
@admin_required
def delete_invoice(inv_id):
    get_supabase().table("invoices").delete().eq("id", inv_id).execute()
    flash("Invoice deleted.", "warning")
    return redirect(url_for("admin.dashboard") + "#pending")


@admin_bp.route("/api/items/search")
@admin_required
def admin_api_items_search():
    q = request.args.get("q", "").strip()
    if len(q) < 1:
        return jsonify([])

    sb = get_supabase()
    results = sb.table("items").select("id, item_code, name, discount").or_(
        f"item_code.ilike.%{q}%,name.ilike.%{q}%"
    ).limit(15).execute().data

    if results:
        item_ids = [r["id"] for r in results]
        stock_rows = sb.table("warehouse_stock").select("item_id, stock").in_("item_id", item_ids).execute().data

        stock_map = {}
        for sr in stock_rows:
            stock_map[sr["item_id"]] = stock_map.get(sr["item_id"], 0) + sr["stock"]

        for r in results:
            r["total_stock"] = stock_map.get(r["id"], 0)

    return jsonify(results)


# ========== STOCK MANAGEMENT ==========
@admin_bp.route("/stock/add", methods=["POST"])
@admin_required
def add_stock():
    warehouse_id = request.form.get("warehouse_id")
    item_id = request.form.get("item_id")
    quantity = request.form.get("quantity", "0").strip() or "0"

    if not warehouse_id or not item_id or int(quantity) <= 0:
        flash("Please fill all fields with valid values.", "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    sb = get_supabase()
    wh_id = int(warehouse_id)
    it_id = int(item_id)
    qty = int(quantity)

    _add_stock_quantity(sb, wh_id, it_id, qty)

    flash(f"Added {qty} units to stock.", "success")
    return redirect(url_for("admin.dashboard") + "#stock")


@admin_bp.route("/stock/upload", methods=["POST"])
@admin_required
def upload_stock_excel():
    warehouse_id = request.form.get("warehouse_id")
    file = request.files.get("stock_file")

    if not warehouse_id:
        flash("Please select a warehouse for stock upload.", "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    if not file or not file.filename:
        flash("Please choose an Excel file to upload.", "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    file_name = file.filename.lower()
    if not file_name.endswith((".xlsx", ".xlsm")):
        flash("Please upload an Excel .xlsx file.", "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    try:
        parsed_rows, issues = _parse_stock_file(file)
    except Exception:
        flash("Could not read the Excel file. Please check the format and try again.", "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    if not parsed_rows:
        flash(issues[0] if issues else "No valid rows found in the uploaded file.", "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    sb = get_supabase()
    wh_id = int(warehouse_id)
    item_rows = sb.table("items").select("id, name, item_code").execute().data

    items_by_name = {str(r.get("name") or "").strip().lower(): r for r in item_rows if r.get("name")}
    items_by_code = {str(r.get("item_code") or "").strip().lower(): r for r in item_rows if r.get("item_code")}

    qty_by_item_id = {}
    missing_items = []
    for row in parsed_rows:
        lookup = row["item_name"].strip().lower()
        item = items_by_name.get(lookup) or items_by_code.get(lookup)
        if not item:
            missing_items.append(row["item_name"])
            continue

        item_id = int(item["id"])
        qty_by_item_id[item_id] = qty_by_item_id.get(item_id, 0) + int(row["quantity"])

    if not qty_by_item_id:
        message = "No matching items found in the uploaded file."
        if missing_items:
            message += f" Missing: {', '.join(sorted(set(missing_items))[:5])}"
        flash(message, "danger")
        return redirect(url_for("admin.dashboard") + "#stock")

    for item_id, qty in qty_by_item_id.items():
        _add_stock_quantity(sb, wh_id, item_id, qty)

    success_message = f"Uploaded stock for {len(qty_by_item_id)} items to the selected warehouse."
    if missing_items:
        success_message += f" Skipped missing items: {', '.join(sorted(set(missing_items))[:5])}"
    elif issues:
        success_message += f" Skipped {len(issues)} invalid row(s)."

    flash(success_message, "success")
    return redirect(url_for("admin.dashboard") + "#stock")
